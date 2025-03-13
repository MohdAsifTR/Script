from flask import Flask, request, render_template, send_file
from pathlib import Path
import uuid
from openpyxl import load_workbook, Workbook
from openpyxl.cell.cell import MergedCell

# Mock constant values for demonstration
class constant:
    TARGET_STRING = "target_string"
    REPLACEMENT_STRING = "replacement_string"
    FILE_SUFFIX = ".xlsx"

app = Flask(__name__)

target = constant.TARGET_STRING
replacement = constant.REPLACEMENT_STRING

def process_cell(cell) -> str:
    value = cell.value
    if cell.data_type == 'f':
        if target in value:
            return value.replace(target, replacement)
    elif isinstance(value, str) and target in value:
        return value.replace(target, replacement)
    return value

def modify_excel_file(src_path: Path, dir_path: Path) -> Path:
    input_wb = load_workbook(src_path)

    for sheet in input_wb.sheetnames:
        input_ws = input_wb[sheet]
        cells = [cell for row in input_ws.iter_rows() for cell in row]
        for cell in cells:
            if not isinstance(cell, MergedCell):
                cell.value = process_cell(cell)

    new_uuid = uuid.uuid4()
    new_filename = f"{new_uuid}{constant.FILE_SUFFIX}"
    new_local_path = dir_path / new_filename
    input_wb.save(new_local_path)

    input_wb.close()
    return new_local_path

@app.route("/", methods=["GET", "POST"])
def upload_file():
    if request.method == "POST":
        file = request.files["file"]
        if file:
            input_path = Path("uploads") / file.filename
            file.save(input_path)
            output_path = modify_excel_file(input_path, Path("uploads"))
            return send_file(output_path, as_attachment=True)
    return render_template("upload.html")

if __name__ == "__main__":
    app.run(debug=True)